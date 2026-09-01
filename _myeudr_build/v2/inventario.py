#!/usr/bin/env python3
"""Inventario dei rilievi ancora aperti: confronta la correzione_proposta col valore
attuale del foglio v2 e classifica (riserva / applicabile / gia' allineata)."""
import glob, json, os, re, sys, unicodedata, collections

HERE  = os.path.dirname(os.path.abspath(__file__))
BUILD = os.path.dirname(HERE); REPO = os.path.dirname(BUILD)
VER   = os.path.join(BUILD, "verifica")
XLSX  = os.environ.get("MYEUDR_XLSX", os.path.join(REPO, "MyEUDR_Lead_Mapping_v2.xlsx"))
COL   = {"denominazione":1,"filiera":2,"dimensione":3,"referente":4,"ruolo":5,
         "linkedin":6,"email":7,"sito":8,"sede":9,"fonte":10}
RISERVA = re.compile(r"DA CONFERMARE|da confermare|riconferm|in alternativa|\?|da verificar|"
                     r"verificare|ipotes|presumib|probabil", re.I)

def fold(x):
    x = (x or "").lower().replace("ø","o").replace("æ","ae").replace("å","a")
    x = unicodedata.normalize("NFKD", x)
    x = "".join(c for c in x if not unicodedata.combining(c))
    x = re.sub(r"[^a-z0-9]", "", re.sub(r"\(.*?\)", "", x))
    return x.replace("oe","o").replace("ae","a").replace("ue","u")

def match(a,b):
    fa,fb = fold(a),fold(b)
    if fa==fb: return True
    if len(fa)>=12 and len(fb)>=12 and (fa.startswith(fb[:12]) or fb.startswith(fa[:12])): return True
    return sorted(fa)==sorted(fb)

def load_rilievi():
    out=[]
    for fp in sorted(glob.glob(os.path.join(VER,"*.json"))):
        b=os.path.basename(fp)
        if b.startswith(("correzioni_","_","00_")): continue
        d=json.load(open(fp,encoding="utf-8"))
        if isinstance(d,dict): d=d.get("rilievi",[])
        for r in d: r["_file"]=b; out.append(r)
    return out

def main():
    from openpyxl import load_workbook
    wb = load_workbook(XLSX)
    idx = {}                       # (foglio, fold(den)) -> row
    for sn in wb.sheetnames:
        ws=wb[sn]
        for row in range(3, ws.max_row+1):
            den=ws.cell(row,1).value
            if den: idx[(sn, fold(den))]=row

    R = load_rilievi()
    out=[]; stat=collections.Counter()
    for r in R:
        prop = (r.get("correzione_proposta") or "").strip()
        camp = (r.get("campo") or "").strip().lower()
        sh   = r.get("foglio")
        if not prop or camp not in COL or sh not in wb.sheetnames:
            stat["senza_proposta_o_campo"]+=1; continue
        ws=wb[sh]; row=idx.get((sh, fold(r.get("denominazione"))))
        if row is None:
            for (s,f),rr in idx.items():
                if s==sh and match(f, r.get("denominazione")): row=rr; break
        if row is None:
            stat["record_non_trovato"]+=1
            out.append(dict(r, _stato="record_non_trovato", _attuale=None, _row=None)); continue
        cur = ws.cell(row, COL[camp]).value
        cur = "" if cur is None else str(cur).strip()
        if fold(cur)==fold(prop) and cur:
            stat["gia_allineata"]+=1; st="gia_allineata"
        elif RISERVA.search(prop):
            stat["con_riserva"]+=1; st="con_riserva"
        else:
            stat["aperta"]+=1; st="aperta"
        out.append(dict(r, _stato=st, _attuale=cur, _row=row))

    json.dump(out, open(os.path.join(HERE,"inventario.json"),"w",encoding="utf-8"),
              ensure_ascii=False, indent=1)
    print("=== STATO ===")
    for k,v in stat.most_common(): print(f"  {k:26s} {v:5d}")
    print("\n=== APERTE per campo x gravita ===")
    c=collections.Counter((x["campo"],x["gravita"]) for x in out if x["_stato"]=="aperta")
    for (camp,g),n in sorted(c.items()): print(f"  {camp:14s} {g:6s} {n:4d}")
    print("\n=== APERTE per foglio ===")
    c=collections.Counter(x["foglio"] for x in out if x["_stato"]=="aperta")
    for k,v in c.most_common(): print(f"  {k:12s} {v:4d}")

if __name__=="__main__": main()

#!/usr/bin/env python3
"""Confronta v2 col backup pre-modifica CONFRONTANDO I RECORD PER IDENTITA', non per
posizione: add_country.py riordina il foglio per filiera e nome, quindi una denominazione
corretta sposta legittimamente la riga. Controlla righe per foglio, ordine dei fogli,
campi svuotati (deve essere ZERO) e ogni cella cambiata."""
import json, os, sys
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
A = load_workbook(os.path.join(HERE, "backup_v1.xlsx"))
B = load_workbook(os.path.join(REPO, "MyEUDR_Lead_Mapping_v2.xlsx"))
ORDINE = ["Italia","Germania","Finlandia","Danimarca","Svezia","Olanda","Belgio","Austria"]
CAMPI = ["denominazione","filiera","dimensione","referente","ruolo","linkedin","email","sito","sede","fonte"]

# rinomine volute: vecchia denominazione -> nuova (dalla tabella delle correzioni manuali)
RINOMINE = {}
for fn in ("correzioni_v2_manuali.json", "correzioni_v2_generiche.json"):
    fp = os.path.join(HERE, fn)
    if not os.path.exists(fp): continue
    for c in json.load(open(fp, encoding="utf-8")):
        if c.get("campo") == "denominazione" and c.get("a"):
            RINOMINE[(c["foglio"], c["denominazione"])] = c["a"]

ok = True
print("ordine fogli v2 :", B.sheetnames)
if B.sheetnames != ORDINE: print("  ✗ ORDINE ERRATO"); ok = False
else: print("  ✓ ordine corretto")

def righe(wb, sn):
    return [[("" if c.value is None else str(c.value).strip()) for c in r]
            for r in wb[sn].iter_rows(min_row=3, max_col=10) if r[0].value]

svuotati, cambi, orfani = [], [], []
print("\nrighe per foglio (v1 → v2):")
for sn in ORDINE:
    ra, rb = righe(A, sn), righe(B, sn)
    flag = "✓" if len(ra) == len(rb) else "✗"
    if flag == "✗": ok = False
    print(f"  {flag} {sn:12s} {len(ra):4d} → {len(rb):4d}")
    idx = {r[0]: r for r in rb}
    for x in ra:
        chiave = RINOMINE.get((sn, x[0]), x[0])
        y = idx.get(chiave)
        if y is None:
            orfani.append(dict(foglio=sn, denominazione=x[0], cercata=chiave)); ok = False; continue
        for i, campo in enumerate(CAMPI):
            if x[i] == y[i]: continue
            rec = dict(foglio=sn, denominazione=y[0], campo=campo, da=x[i], a=y[i])
            cambi.append(rec)
            if x[i] and x[i].lower() not in ("n.d.", "") and (not y[i] or y[i].lower() == "n.d."):
                svuotati.append(rec)

print(f"\ncelle cambiate : {len(cambi)}")
print(f"record orfani  : {len(orfani)}", "✓" if not orfani else "✗ un record di v1 non si ritrova in v2")
for o in orfani: print("   ", o)
print(f"campi SVUOTATI : {len(svuotati)}", "✓" if not svuotati else "✗ REGOLA 2 VIOLATA")
for s in svuotati: print("   ", s)
json.dump(cambi, open(os.path.join(HERE, "diff_v1_v2.json"), "w", encoding="utf-8"),
          ensure_ascii=False, indent=1)
if not ok or svuotati: sys.exit(1)
print("\n✓ integrita' verificata")

#!/usr/bin/env python3
"""
Applica le correzioni della v2 a MyEUDR_Lead_Mapping_v2.xlsx.
L'originale MyEUDR_Lead_Mapping.xlsx non viene mai toccato.

Riprende le due regole di sicurezza gia' collaudate:
  · GUARDIA sul valore attuale: se non coincide con quello atteso, la correzione e' saltata
    (lo script resta cosi' rieseguibile senza rischi);
  · NON svuota mai un campo gia' valorizzato.

E i due percorsi, perche' i fogli non hanno tutti la stessa origine:
  · DK/SE/NL/BE/AT hanno i JSON di build -> si corregge il JSON e si rigenera con add_country.py;
  · IT/DE/FI esistono solo come foglio -> si corregge la cella in posto.

Uso:  python applica_v2.py [--dry-run]
"""
import glob, json, os, re, subprocess, sys, unicodedata

HERE  = os.path.dirname(os.path.abspath(__file__))
BUILD = os.path.dirname(HERE); REPO = os.path.dirname(BUILD)
XLSX  = os.environ.get("MYEUDR_XLSX", os.path.join(REPO, "MyEUDR_Lead_Mapping_v2.xlsx"))
ORDINE  = ["Italia","Germania","Finlandia","Danimarca","Svezia","Olanda","Belgio","Austria"]
PREFIX  = {"Danimarca":"dk","Svezia":"se","Olanda":"nl","Belgio":"be","Austria":"at"}
INPLACE = {"Italia","Germania","Finlandia"}
COL = {"denominazione":1,"filiera":2,"dimensione":3,"referente":4,"ruolo":5,
       "linkedin":6,"email":7,"sito":8,"sede":9,"fonte":10}
DRY = "--dry-run" in sys.argv

def fold(x):
    x = (x or "").lower().replace("ø","o").replace("æ","ae").replace("å","a")
    x = unicodedata.normalize("NFKD", x)
    x = "".join(c for c in x if not unicodedata.combining(c))
    x = re.sub(r"[^a-z0-9]", "", re.sub(r"\(.*?\)", "", x))
    return x.replace("oe","o").replace("ae","a").replace("ue","u")

def match(a, b, esatto=False):
    """Confronto fra denominazioni.
    Il ripiego fuzzy e' volutamente stretto: la vecchia regola 'primi 12 caratteri'
    faceva collidere fra loro TUTTE le societa' svedesi che cominciano per
    'Aktiebolaget' (sono sei nel foglio). Bug scoperto in v2 e intercettato dalla
    guardia sul valore attuale."""
    fa, fb = fold(a), fold(b)
    if fa == fb: return True
    if esatto: return False
    corto, lungo = (fa, fb) if len(fa) <= len(fb) else (fb, fa)
    if len(corto) >= 18 and lungo.startswith(corto): return True   # prefisso distintivo
    return len(fa) >= 18 and sorted(fa) == sorted(fb)              # termini invertiti

S = dict(app=0, salt=0, gia=0, nf=0)
LOG = []

def registra(sh, den, campo, da, a, c, esito):
    LOG.append(dict(foglio=sh, denominazione=den, campo=campo, da=da, a=a, esito=esito,
                    motivo=c.get("motivo",""), fonte=c.get("fonte",""),
                    gravita=c.get("gravita",""), origine=c.get("origine","")))

def campi_di(c):
    """(campo, valore) da applicare: tabella generica o tabella referenti."""
    if "campo" in c: return [(c["campo"], c["a"])]
    return [(k, c.get(k)) for k in ("referente","ruolo") if (c.get(k) or "").strip()]

def applica_a_record(get, set_, c, sh):
    """get(campo)->valore attuale ; set_(campo, valore). Torna True se ha toccato qualcosa."""
    tocco = False
    for campo, val in campi_di(c):
        cur = get(campo); cur = "" if cur is None else str(cur).strip()
        val = "" if val is None else str(val)
        if not val.strip():                       # regola 2: mai svuotare un campo
            continue

        # __APPEND__ = aggiunge una clausola in coda invece di sostituire il campo.
        # Non avendo una guardia "da", pretende un'ancora: un frammento che DEVE
        # comparire nel valore attuale, cosi' la coda non puo' finire su un altro record.
        if val.startswith("__APPEND__"):
            coda = val[len("__APPEND__"):].strip()
            if coda.lower() in cur.lower():
                S["gia"] += 1
                registra(sh, c["denominazione"], campo, cur, coda, c, "gia_presente"); continue
            anc = (c.get("ancora") or "").strip()
            if not anc or anc.lower() not in cur.lower():
                S["salt"] += 1
                registra(sh, c["denominazione"], campo, cur, coda, c, "saltata_ancora")
                print(f"  ~ {sh}/{c['denominazione'][:30]} {campo}: ancora «{anc[:34]}» assente — SALTATA")
                continue
            val = (cur + " " + coda).strip() if cur and cur.lower() != "n.d." else coda

        if fold(cur) == fold(val):
            S["gia"] += 1; registra(sh, c["denominazione"], campo, cur, val, c, "gia_uguale"); continue
        atteso = c.get("da")
        if "campo" in c and atteso is not None and cur != str(atteso).strip():
            S["salt"] += 1; registra(sh, c["denominazione"], campo, cur, val, c, "saltata_guardia")
            print(f"  ~ {sh}/{c['denominazione'][:30]} {campo}: attuale «{cur[:40]}» ≠ atteso «{str(atteso)[:40]}» — SALTATA")
            continue
        if not DRY: set_(campo, val)
        S["app"] += 1; tocco = True
        registra(sh, c["denominazione"], campo, cur, val, c, "applicata")
        print(f"  ✔ {sh}/{c['denominazione'][:28]:30s} {campo:13s} «{cur[:34]}» → «{str(val)[:40]}»")
    return tocco

def main():
    from openpyxl import load_workbook
    fp_rim = os.path.join(HERE, "rimozioni_v2.json")
    RIMOSSI = {(r["foglio"], r["denominazione"]) for r in
               (json.load(open(fp_rim, encoding="utf-8")) if os.path.exists(fp_rim) else [])}
    C = []
    for nome in ("correzioni_v2_generiche","correzioni_v2_referenti","correzioni_v2_manuali",
                 "correzioni_v2_agenti","correzioni_v2_agenti_ref",
                 "correzioni_v2_dimensione", "correzioni_v2_sede",
                 "correzioni_v2_residui", "correzioni_v2_gruppi",
                 "correzioni_v2_passo4", "correzioni_v2_taglia_sopra",
                 "correzioni_v2_taglia_sotto"):
        fp = os.path.join(HERE, nome + ".json")
        if os.path.exists(fp): C += json.load(open(fp, encoding="utf-8"))
    # una correzione chiave sul nome VECCHIO non troverebbe piu' un record rinominato:
    # allineo le chiavi alle rinomine gia' in tabella, cosi' la riesecuzione resta pulita
    RINOMINE = {(c["foglio"], c["denominazione"]): c["a"]
                for c in C if c.get("campo") == "denominazione" and c.get("a")}
    for c in C:
        nuovo = RINOMINE.get((c["foglio"], c["denominazione"]))
        if nuovo and c.get("campo") != "denominazione":
            c["denominazione"] = nuovo
    print(f"{len(C)} correzioni in tabella · target {os.path.basename(XLSX)}\n")

    # --- 1. fogli con JSON di build ---
    tocchi = set()
    for c in C:
        sh = c["foglio"]
        if sh in INPLACE: continue
        cand = []
        for fp in sorted(glob.glob(os.path.join(BUILD, f"{PREFIX[sh]}_*.json"))):
            data = json.load(open(fp, encoding="utf-8"))
            for r in data:
                if match(r.get("denominazione"), c["denominazione"], esatto=True):
                    cand.append((fp, data, r, True))
                elif match(r.get("denominazione"), c["denominazione"]):
                    cand.append((fp, data, r, False))
        esatti = [x for x in cand if x[3]] or cand      # l'esatto batte sempre il fuzzy
        trovato = len(esatti) == 1
        if len(esatti) > 1:
            S["salt"] += 1
            print(f"  ~ {sh}: «{c['denominazione']}» corrisponde a {len(esatti)} record — AMBIGUA, saltata")
        elif trovato:
            fp, data, r, _ = esatti[0]
            if applica_a_record(lambda k: r.get(k), lambda k,v: r.__setitem__(k,v), c, sh):
                if not DRY:
                    json.dump(data, open(fp,"w",encoding="utf-8"), ensure_ascii=False, indent=1)
                tocchi.add(sh)
        if not trovato and len(esatti) <= 1:
            if c.get("campo") == "denominazione" and any(
                    match(r.get("denominazione"), c["a"], esatto=True)
                    for fp in glob.glob(os.path.join(BUILD, f"{PREFIX[sh]}_*.json"))
                    for r in json.load(open(fp, encoding="utf-8"))):
                S["gia"] += 1; print(f"  = {sh}: «{c['denominazione']}» gia' rinominata")
            elif (sh, c["denominazione"]) in RIMOSSI:
                print(f"  = {sh}: «{c['denominazione']}» rimossa dal censimento")
            else:
                S["nf"] += 1; print(f"  !! {sh}: «{c['denominazione']}» non trovata nei JSON")

    for sh in sorted(tocchi):
        if DRY: print(f"  (dry-run) rigenererei {sh}"); continue
        subprocess.run([sys.executable, os.path.join(BUILD,"add_country.py"), PREFIX[sh], sh, XLSX],
                       check=True, capture_output=True)
        print(f"  ↻ foglio {sh} rigenerato")

    # --- 2. fogli senza JSON: cella in posto ---
    wb = load_workbook(XLSX); changed = False
    for c in C:
        sh = c["foglio"]
        if sh not in INPLACE: continue
        ws = wb[sh]
        cand = [row for row in range(3, ws.max_row+1)
                if ws.cell(row,1).value and match(ws.cell(row,1).value, c["denominazione"], esatto=True)]
        if not cand:
            cand = [row for row in range(3, ws.max_row+1)
                    if ws.cell(row,1).value and match(ws.cell(row,1).value, c["denominazione"])]
        trovato = len(cand) == 1
        if len(cand) > 1:
            S["salt"] += 1
            print(f"  ~ {sh}: «{c['denominazione']}» corrisponde a {len(cand)} righe — AMBIGUA, saltata")
        elif trovato:
            row = cand[0]
            if applica_a_record(lambda k: ws.cell(row, COL[k]).value,
                                lambda k,v: setattr(ws.cell(row, COL[k]), "value", v), c, sh):
                changed = True
        if not trovato and len(cand) <= 1:
            if c.get("campo") == "denominazione" and any(
                    ws.cell(rr,1).value and match(ws.cell(rr,1).value, c["a"], esatto=True)
                    for rr in range(3, ws.max_row+1)):
                S["gia"] += 1; print(f"  = {sh}: «{c['denominazione']}» gia' rinominata")
            else:
                S["nf"] += 1; print(f"  !! {sh}: «{c['denominazione']}» non trovata nel foglio")

    # --- 3. ordine dei fogli + salvataggio ---
    if not DRY and (changed or tocchi):
        wb._sheets = [wb[n] for n in ORDINE]
        wb.save(XLSX)
        print(f"\n↻ ordine fogli: {load_workbook(XLSX).sheetnames}")
    json.dump(LOG, open(os.path.join(HERE,"log_applicazione.json"),"w",encoding="utf-8"),
              ensure_ascii=False, indent=1)
    print(f"\napplicate {S['app']} · gia uguali {S['gia']} · saltate dalla guardia {S['salt']} · "
          f"non trovate {S['nf']}" + ("  (DRY-RUN)" if DRY else ""))

if __name__ == "__main__": main()

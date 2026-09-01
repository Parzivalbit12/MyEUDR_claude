#!/usr/bin/env python3
"""Rimozioni dalla v2. Il mandato le autorizza SOLO se (a) la commodity e' fuori
Allegato I, oppure (b) l'azienda non esiste piu' o non produce piu'.
Il fuori taglia da solo NON basta e non e' mai motivo di rimozione."""
import glob, json, os, subprocess, sys, unicodedata, re

HERE = os.path.dirname(os.path.abspath(__file__))
BUILD = os.path.dirname(HERE); REPO = os.path.dirname(BUILD)
XLSX = os.path.join(REPO, "MyEUDR_Lead_Mapping_v2.xlsx")
ORDINE = ["Italia","Germania","Finlandia","Danimarca","Svezia","Olanda","Belgio","Austria"]
PREFIX = {"Danimarca":"dk","Svezia":"se","Olanda":"nl","Belgio":"be","Austria":"at"}

RIMOZIONI = [dict(
    foglio="Danimarca", denominazione="Skagerak Denmark A/S", categoria="b",
    motivo="Societa' estinta: il CVR 28855990 risulta «opløst efter fusion» e non ha piu' "
           "alcun direktør registrato. La persona giuridica censita non esiste piu'. "
           "Stesso criterio gia' applicato in v1 a Odense Seglmærkefabrik («opløst efter fusion») "
           "e a Kaffekompaniet. Si aggiunge il legame gia' noto con Fritz Hansen, fra i big esclusi.",
    fonte="https://estatistik.dk/virksomhed/skagerak-denmark-as/28855990/roller — «0 direktører; status: OPLØST EFTER FUSION»",
), dict(
    foglio="Svezia", denominazione="Lenanders Grafiska AB", categoria="b",
    motivo="Non produce piu'. A meta' 2021 Scandinavian Print Group ha rilevato clienti e "
           "contratti e soltanto sette dipendenti; oltre venti addetti sono stati licenziati e "
           "la PRODUZIONE di Kalmar e' stata chiusa. Il marchio Lenanders sopravvive come "
           "insegna del gruppo, ma l'org.nr storico 556592-3066 risulta ridenominato e "
           "cancellato dai registri fiscali. Stesso criterio gia' applicato in v1 a Bayer "
           "Kartonagen (asset deal, produzione trasferita) e a Marandi (produzione cessata).",
    fonte="Riverifica lotto E — «Scandinavian Print Group took over customers, contracts, and a "
          "total of seven employees. Over 20 employees at the printing facility were laid off "
          "due to the acquisition. The production in Kalmar was discontinued.»",
)]

def fold(x):
    x = (x or "").lower().replace("ø","o").replace("æ","ae").replace("å","a")
    x = unicodedata.normalize("NFKD", x)
    x = "".join(c for c in x if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", re.sub(r"\(.*?\)", "", x))

def main():
    from openpyxl import load_workbook
    fatte = []
    for R in RIMOZIONI:
        sh = R["foglio"]
        if sh not in PREFIX:
            print(f"  !! {sh}: foglio senza JSON di build, rimozione non gestita"); continue
        tolto = False
        for fp in sorted(glob.glob(os.path.join(BUILD, f"{PREFIX[sh]}_*.json"))):
            data = json.load(open(fp, encoding="utf-8"))
            keep = [r for r in data if fold(r.get("denominazione")) != fold(R["denominazione"])]
            if len(keep) == len(data): continue
            json.dump(keep, open(fp, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
            print(f"  - {sh}: rimosso «{R['denominazione']}» da {os.path.basename(fp)}")
            tolto = True; fatte.append(R); break
        if not tolto:
            print(f"  = {sh}: «{R['denominazione']}» gia' assente")
    if fatte:
        for sh in {R["foglio"] for R in fatte}:
            subprocess.run([sys.executable, os.path.join(BUILD,"add_country.py"), PREFIX[sh], sh, XLSX],
                           check=True, capture_output=True)
            print(f"  ↻ foglio {sh} rigenerato")
        wb = load_workbook(XLSX); wb._sheets = [wb[n] for n in ORDINE]; wb.save(XLSX)
        print(f"  ↻ ordine fogli: {load_workbook(XLSX).sheetnames}")
    json.dump(RIMOZIONI, open(os.path.join(HERE,"rimozioni_v2.json"),"w",encoding="utf-8"),
              ensure_ascii=False, indent=1)

if __name__ == "__main__": main()

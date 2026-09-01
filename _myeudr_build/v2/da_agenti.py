#!/usr/bin/env python3
"""Trasforma gli esiti degli agenti di riverifica (PASSO 3) in tabelle di correzione,
leggendo il valore attuale dal foglio v2 per costruire la guardia 'da'.
Applica solo gli esiti 'risolto' che portano un valore e una fonte."""
import glob, json, os, re, unicodedata
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(os.path.dirname(HERE))
COL = {"denominazione":1,"filiera":2,"dimensione":3,"referente":4,"ruolo":5,
       "linkedin":6,"email":7,"sito":8,"sede":9,"fonte":10}

def fold(x):
    x = (x or "").lower().replace("ø","o").replace("æ","ae").replace("å","a")
    x = unicodedata.normalize("NFKD", x)
    x = "".join(c for c in x if not unicodedata.combining(c))
    x = re.sub(r"[^a-z0-9]", "", re.sub(r"\(.*?\)", "", x))
    return x.replace("oe","o").replace("ae","a").replace("ue","u")

wb = load_workbook(os.path.join(REPO, "MyEUDR_Lead_Mapping_v2.xlsx"))
IDX = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in range(3, ws.max_row+1):
        d = ws.cell(row,1).value
        if d: IDX.setdefault((sn, fold(d)), []).append(row)

gen, ref, scarti = [], [], []
import sys
# solo gli agenti indicati: un file di output ancora in scrittura (salvataggio
# incrementale) non va consumato finche' l'agente non ha finito
quali = sys.argv[1:] or ["*"]
FILES = sorted({f for q in quali for f in glob.glob(os.path.join(HERE, "riverifica", f"out_{q}*.json"))})
print("consumo:", [os.path.basename(f) for f in FILES])
for fp in FILES:
    for x in json.load(open(fp, encoding="utf-8")):
        sh, den = x.get("foglio"), x.get("denominazione")
        if x.get("esito") != "risolto":
            scarti.append(dict(x, _perche=f"esito «{x.get('esito')}»: nessuna modifica")); continue
        righe = IDX.get((sh, fold(den)), [])
        if len(righe) != 1:
            scarti.append(dict(x, _perche=f"{len(righe)} righe corrispondono: non applicata")); continue
        ws = wb[sh]; row = righe[0]
        base = dict(foglio=sh, denominazione=ws.cell(row,1).value,
                    motivo=x.get("nota",""), fonte=x.get("fonte","") + " — " + x.get("frammento","")[:200],
                    gravita="", origine=os.path.basename(fp))
        if "campo" in x:                                     # agente B: un campo per rilievo
            campo, val = x["campo"], (x.get("valore") or "").strip()
            if not val or campo not in COL:
                scarti.append(dict(x, _perche="valore vuoto o campo ignoto")); continue
            cur = ws.cell(row, COL[campo]).value
            gen.append(dict(base, campo=campo, da="" if cur is None else str(cur).strip(), a=val))
        elif "dimensione" in x:                               # agente C
            val = (x.get("dimensione") or "").strip()
            if not val: scarti.append(dict(x, _perche="dimensione vuota")); continue
            cur = ws.cell(row, COL["dimensione"]).value
            gen.append(dict(base, campo="dimensione", da="" if cur is None else str(cur).strip(), a=val))
        else:                                                 # agente A: referente/ruolo
            r, ru = (x.get("referente") or "").strip(), (x.get("ruolo") or "").strip()
            if not r and not ru: scarti.append(dict(x, _perche="ne' referente ne' ruolo")); continue
            ref.append(dict(base, referente=r, ruolo=ru))

json.dump(gen, open(os.path.join(HERE,"correzioni_v2_agenti.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)
json.dump(ref, open(os.path.join(HERE,"correzioni_v2_agenti_ref.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)
json.dump(scarti, open(os.path.join(HERE,"agenti_non_applicati.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"generiche {len(gen)} · referenti {len(ref)} · non applicati {len(scarti)}")

#!/usr/bin/env python3
"""
Aggiunge (o sostituisce) UN foglio-paese nel workbook MyEUDR, lasciando intatti gli altri.

Uso:
    python add_country.py <prefix> <NomeFoglio> <percorso_xlsx>
Esempio:
    python add_country.py dk Danimarca "C:/Users/39333/Desktop/EDUR da Studio/MyEUDR_Lead_Mapping.xlsx"

Legge tutti i file <prefix>_*.json (array di oggetti-lead) nella cartella di questo script,
normalizza le entità HTML, deduplica per nome (ignorando le parentesi), ordina per filiera,
poi crea/sostituisce il foglio <NomeFoglio> nel workbook. Se il workbook non esiste lo crea.

Schema di ogni oggetto JSON:
  denominazione, filiera, sede, sito, email, referente, ruolo, linkedin, dimensione, fonte
"""
import glob, html, json, os, re, sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("denominazione", "Denominazione", 34),
    ("filiera",       "Filiera EUDR", 20),
    ("dimensione",    "Dimensione (fatt./dip.)", 30),
    ("referente",     "Referente", 22),
    ("ruolo",         "Ruolo", 20),
    ("linkedin",      "LinkedIn", 34),
    ("email",         "Email / PEC", 30),
    ("sito",          "Sito web", 30),
    ("sede",          "Sede", 22),
    ("fonte",         "Fonte", 34),
]
HEADER_FILL = PatternFill("solid", fgColor="1F6B3B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color="1F6B3B", size=13)
ALT_FILL = PatternFill("solid", fgColor="EEF5EF")
THIN = Side(style="thin", color="D0D7D2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)
ORDER = ["pelle","leder","læder","legno","holz","tr", "wood","carta","papier","papir",
         "caffè","caffe","kaffee","kaffe","cacao","kakao","gomma","kautschuk","gummi",
         "mangimi","soia","soja","foder","futter","palma","palme","palmöl","bovini",
         "carne","fleisch","kød","oksek"]

def norm_name(s):
    s = re.sub(r"\(.*?\)", "", s or "")          # ignora le parentesi: "X (Y)" == "X"
    return re.sub(r"[^a-z0-9]", "", s.lower())

def clean(v):
    return html.unescape(v) if isinstance(v, str) else v

def fgroup(r):
    f = (r.get("filiera") or "").lower()
    for i, k in enumerate(ORDER):
        if k in f:
            return i
    return 99

def build_sheet(ws, rows):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    t = ws.cell(1, 1, f"MyEUDR \u2014 Lead mapping \u00b7 {ws.title} \u00b7 aziende potenzialmente soggette all'EUDR")
    t.font = TITLE_FONT; t.alignment = Alignment(vertical="center"); ws.row_dimensions[1].height = 22
    for ci, (_, label, w) in enumerate(COLUMNS, 1):
        c = ws.cell(2, ci, label); c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True); c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 26
    for ri, raw in enumerate(rows, 3):
        rec = {k: (clean(raw.get(k)) or "").strip() for k, _, _ in COLUMNS}
        for k in ("dimensione", "email"):
            if not rec[k]:
                rec[k] = "n.d."
        for ci, (k, _, _) in enumerate(COLUMNS, 1):
            c = ws.cell(ri, ci, rec[k]); c.alignment = WRAP_TOP; c.border = BORDER
            if (ri - 2) % 2 == 0:
                c.fill = ALT_FILL
        for k, ci in (("sito", 8), ("linkedin", 6)):
            if rec[k].startswith("http"):
                cell = ws.cell(ri, ci); cell.hyperlink = rec[k]
                cell.font = Font(color="1155CC", underline="single")
        em = rec["email"]
        if em and em != "n.d." and "@" in em and " " not in em:
            cell = ws.cell(ri, 7); cell.hyperlink = f"mailto:{em}"
            cell.font = Font(color="1155CC", underline="single")
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{2 + max(len(rows), 1)}"

def main():
    prefix, name, xlsx = sys.argv[1], sys.argv[2], sys.argv[3]
    here = os.path.dirname(os.path.abspath(__file__))
    rows = []
    for fp in sorted(glob.glob(os.path.join(here, f"{prefix}_*.json"))):
        with open(fp, encoding="utf-8") as f:
            rows += json.load(f)
        print("letto", os.path.basename(fp))
    rows = [{k: clean(v) for k, v in r.items()} for r in rows]
    seen, dd = set(), []
    for r in rows:
        k = norm_name(r.get("denominazione"))
        if k in seen:
            continue
        seen.add(k); dd.append(r)
    dd.sort(key=lambda r: (fgroup(r), r.get("denominazione", "")))
    wb = load_workbook(xlsx) if os.path.exists(xlsx) else Workbook()
    if not os.path.exists(xlsx) and "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    if name in wb.sheetnames:
        wb.remove(wb[name])
    ws = wb.create_sheet(title=name[:31])
    build_sheet(ws, dd)
    wb.save(xlsx)
    em = sum(1 for r in dd if r.get("email") and r["email"] != "n.d.")
    rf = sum(1 for r in dd if (r.get("referente") or "").strip())
    li = sum(1 for r in dd if (r.get("linkedin") or "").strip())
    print(f"OK {name}: {len(dd)} aziende (email {em}/{len(dd)} | referente {rf} | LinkedIn {li}) -> {xlsx}")

if __name__ == "__main__":
    main()

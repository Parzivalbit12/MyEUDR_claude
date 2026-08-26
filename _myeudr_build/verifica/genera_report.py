#!/usr/bin/env python3
"""Genera REPORT_VERIFICA.md: sintesi Fase A + Fase B, casi alta, correzioni proposte."""
import glob, json, os, re
from collections import Counter, defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
SHEETS = ["Italia","Germania","Finlandia","Danimarca","Svezia","Olanda","Belgio","Austria"]
def _conta_fogli():
    """conta le righe reali del workbook: dopo le rimozioni il totale non e' piu' 742"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(os.path.join(os.path.dirname(os.path.dirname(HERE)),
                                        "MyEUDR_Lead_Mapping.xlsx"), read_only=True)
        return {sn: sum(1 for r in wb[sn].iter_rows(min_row=3, values_only=True) if r and r[0])
                for sn in SHEETS}
    except Exception:
        return dict(zip(SHEETS, [95,97,84,89,89,100,95,93]))
ATTESI = _conta_fogli()
NTOT = sum(ATTESI.values())
GRAV = ["alta","media","bassa"]
SKIP = ("_records", "correzioni_")

def esc(s): return str(s or "").replace("|", "\\|").replace("\n", " ").strip()

def carica_rilievi():
    out, files = [], []
    for fp in sorted(glob.glob(os.path.join(HERE, "*.json"))):
        bn = os.path.basename(fp)
        if bn.startswith(SKIP): continue
        try: data = json.load(open(fp, encoding="utf-8"))
        except Exception: continue
        if not isinstance(data, list): continue
        files.append(bn)
        for o in data:
            if not isinstance(o, dict): continue
            o["_file"] = bn
            g = (o.get("gravita") or "").strip().lower()
            o["gravita"] = g if g in GRAV else "media"
            out.append(o)
    return out, files

def blocchi_coperti(files):
    """quali blocchi di ciascun foglio sono stati verificati, distinguendo completi e parziali"""
    try:
        stato = json.load(open(os.path.join(HERE, "_stato_blocchi.json"), encoding="utf-8"))
    except Exception:
        stato = {}
    tot = defaultdict(list); fatti = defaultdict(list); parz = defaultdict(list)
    for fp in sorted(glob.glob(os.path.join(HERE, "blocchi", "*.json"))):
        bn = os.path.basename(fp)[:-5]
        sh = bn.rsplit("_", 1)[0].capitalize()
        n = len(json.load(open(fp, encoding="utf-8")))
        tot[sh].append((bn, n))
        if bn + ".json" not in files: continue
        st = stato.get(bn, {})
        if st.get("stato") == "completo":
            fatti[sh].append((bn, int(st.get("verificati") or n)))
        else:
            # blocco non confermato completo (in corso o interrotto): conto come copertura
            # solo le aziende che compaiono davvero fra i rilievi — è un minimo garantito
            try:
                d = json.load(open(os.path.join(HERE, bn + ".json"), encoding="utf-8"))
                v = len({str(o.get("denominazione", "")) for o in d if isinstance(o, dict)})
            except Exception:
                v = int(st.get("verificati") or 0)
            parz[sh].append((bn, v))
    return tot, fatti, parz

def main():
    R, files = carica_rilievi()
    tot_b, fatti_b, parz_b = blocchi_coperti(files)
    by_sheet = defaultdict(list)
    for r in R: by_sheet[r.get("foglio", "?")].append(r)
    c_all = Counter(r["gravita"] for r in R)

    o = []
    o.append("# REPORT DI VERIFICA — MyEUDR Lead Mapping\n")
    o.append(f"Controllo qualità record per record del censimento lead (**{NTOT} aziende, 8 fogli**). "
             "Non è una raccolta di nuove aziende: è la verifica del lavoro esistente.\n")
    o.append("La verifica si è svolta in due fasi:\n")
    o.append("- **Fase A — controlli deterministici**, offline, su tutti i JSON di build e sul "
             "workbook: 24 controlli automatici su duplicati, email dedotte, URL, entità HTML, "
             "tassonomia, forbice dimensionale, denominazioni e forme giuridiche. "
             "Dettaglio in [`00_controlli_automatici.md`](00_controlli_automatici.md) e "
             "[`01_analisi_dimensione.md`](01_analisi_dimensione.md).")
    o.append("- **Fase B — riscontro sul web**, record per record, tramite agenti di verifica "
             "che hanno lavorato a blocchi di 15-20 aziende con 2-3 ricerche ciascuna. "
             "Ogni rilievo porta un URL o la citazione del frammento a sostegno.\n")

    # ---------- copertura ----------
    o.append("\n## 1. Copertura della verifica\n")
    o.append("| Foglio | Aziende | Blocchi completi | Blocchi parziali | Blocchi da fare | Aziende verificate | Copertura |")
    o.append("|---|--:|--:|--:|--:|--:|--:|")
    tot_ver = 0
    for sn in SHEETS:
        nb, nf, npz = tot_b.get(sn, []), fatti_b.get(sn, []), parz_b.get(sn, [])
        nrec = sum(n for _, n in nf) + sum(n for _, n in npz); tot_ver += nrec
        o.append(f"| {sn} | {ATTESI[sn]} | {len(nf)} | {len(npz)} | {len(nb)-len(nf)-len(npz)} | "
                 f"{nrec} | {100*nrec/ATTESI[sn]:.0f}% |")
    nfa = sum(len(v) for v in fatti_b.values()); npa = sum(len(v) for v in parz_b.values())
    nto = sum(len(v) for v in tot_b.values())
    o.append(f"| **TOTALE** | **{NTOT}** | **{nfa}** | **{npa}** | **{nto-nfa-npa}** | "
             f"**{tot_ver}** | **{100*tot_ver/NTOT:.0f}%** |")
    o.append("\n_Un blocco è contato **completo** solo se l'agente ha confermato di aver verificato "
             "tutti i record. I **blocchi parziali** sono quelli ancora in corso o interrotti dal "
             "limite di sessione: i rilievi già salvati sono validi e inclusi nel report, ma la "
             "copertura è conteggiata al ribasso (solo le aziende che compaiono fra i rilievi). "
             "Il salvataggio incrementale ogni 3-4 record è ciò che ha evitato di perdere quel lavoro._")
    o.append(f"\n> La Fase A copre invece il **100%** dei {NTOT} record: è un controllo offline "
             "e non dipende dal budget di ricerca.\n")
    o.append("\n_A questi si aggiunge la verifica mirata dei **13 punti già noti** lasciati aperti "
             "dalla raccolta, condotta separatamente e riportata per intero più sotto._\n")

    # ---------- sintesi rilievi ----------
    o.append("\n## 2. Rilievi per foglio\n")
    o.append(f"**Totale rilievi Fase B: {len(R)}** — alta {c_all['alta']} · "
             f"media {c_all['media']} · bassa {c_all['bassa']}.\n")
    o.append("| Foglio | Rilievi | alta | media | bassa | Aziende toccate |")
    o.append("|---|--:|--:|--:|--:|--:|")
    for sn in SHEETS:
        rs = by_sheet.get(sn, [])
        c = Counter(r["gravita"] for r in rs)
        o.append(f"| {sn} | {len(rs)} | {c['alta']} | {c['media']} | {c['bassa']} | "
                 f"{len({r.get('denominazione','') for r in rs})} |")
    altri = [k for k in by_sheet if k not in SHEETS]
    for sn in altri:
        rs = by_sheet[sn]; c = Counter(r["gravita"] for r in rs)
        o.append(f"| _{sn}_ | {len(rs)} | {c['alta']} | {c['media']} | {c['bassa']} | "
                 f"{len({r.get('denominazione','') for r in rs})} |")
    o.append(f"| **TOTALE** | **{len(R)}** | **{c_all['alta']}** | **{c_all['media']}** | "
             f"**{c_all['bassa']}** | **{len({(r.get('foglio'),r.get('denominazione')) for r in R})}** |")

    o.append("\n### Rilievi per campo\n")
    fc = Counter(str(r.get("campo","?")).strip().lower() for r in R)
    fa = Counter(str(r.get("campo","?")).strip().lower() for r in R if r["gravita"]=="alta")
    o.append("| Campo | Rilievi | di cui alta |")
    o.append("|---|--:|--:|")
    for f, n in fc.most_common():
        o.append(f"| {esc(f)} | {n} | {fa[f]} |")

    # ---------- tema trasversale: legami di gruppo ----------
    recs = {(r["_sheet"], r["denominazione"]): r
            for r in json.load(open(os.path.join(HERE, "_records.json"), encoding="utf-8"))}
    GRP = re.compile(r"controllat|capogruppo|gruppo|koncern|datterselskab|dotterbolag|moderbolag|"
                     r"onderdeel van|dochter|Tochter|Konzern|part of|acquisit|maggioranza di|"
                     r"holding|ejet af|ägs av", re.I)
    # solo rilievi il cui MERITO è l'assetto proprietario, non quelli che citano "gruppo" di sfuggita
    OWN = re.compile(r"non (?:e'|è) (?:piu' |più )?indipendent|lead non indipendent|"
                     r"controllata (?:di|da|integrale|del)|controllat[oa] (?:dal|dalla)|"
                     r"capogruppo (?:sbagliat|errat|indicata|non dichiarat|manca)|"
                     r"appartenenza di gruppo|legame di gruppo|legame non dichiarat|"
                     r"assetto proprietario|moderbolag|dotterbolag non dichiarat|"
                     r"filiale di gruppo|controllata di gruppo|societa' non indipendente|"
                     r"società non indipendente|acquisita da|rilevata da|"
                     r"fa parte (?:del|della|di) grupp|nel portafoglio del fondo|"
                     r"socio unico|partecipata (?:da|dal)|koncernmoderbolag", re.I)
    def is_own(r):
        t = str(r.get("problema", ""))
        if not OWN.search(t): return False
        # esclude i rilievi in cui l'assetto è solo contesto di un rilievo su altro
        if re.match(r"^\s*(IDENTITA|Dato obsolet|Discordanza|Referente e ruolo errati|"
                    r"nessuna delle 7)", t, re.I): return False
        return True
    gr = [r for r in R if is_own(r)]
    if gr:
        o.append(f"\n---\n\n## 3. Tema trasversale — legami di gruppo ({len(gr)} rilievi)\n")
        o.append("È il problema **più diffuso e meno atteso** emerso dalla verifica: non era fra i "
                 "13 punti noti dell'handoff. Numerose aziende del censimento sono controllate di "
                 "gruppi, spesso esteri o quotati. Per il criterio già applicato dal progetto — che "
                 "aveva rimosso Lavazza Kaffee, Segafredo Zanetti Austria e Kaffee Partner Austria "
                 "perché *«la compliance si decide a livello di gruppo, non nella filiale»* — sono "
                 "**lead di valore dubbio**.\n")
        o.append("La tabella distingue i due casi, che non hanno la stessa gravità:\n")
        o.append("- **DICHIARATO** — il campo `Dimensione` del foglio già segnala il legame. "
                 "Non è un errore di dato: la raccolta ha fatto quel che le regole chiedevano "
                 "(*«segnalare sempre i legami di gruppo»*). È una **decisione di selezione** "
                 "che spetta al cliente.\n")
        o.append("- **NON DICHIARATO / ERRATO** — il legame manca del tutto, oppure la capogruppo "
                 "indicata è sbagliata. Questo **è** un errore di dato.\n")
        o.append("\n> ⚠️ **Non tutti i legami di gruppo pesano allo stesso modo — leggere questa "
                 "tabella con questo filtro.** Vanno distinti due casi che il conteggio grezzo "
                 "confonde:\n>\n"
                 "> - **Holding di proprietà** — la società è posseduta dal veicolo dei suoi stessi "
                 "soci (`X Holding AB`, `X Förvaltning AB`, `X Invest`). In Svezia e Danimarca è la "
                 "struttura **normale** di un'impresa familiare, e in Olanda il *directeur* "
                 "statutario è spesso una holding proprio per questo. **Non sposta la decisione di "
                 "compliance fuori dall'azienda**: il lead resta valido. Rientrano qui, per esempio, "
                 "Sunnerbo Fönster, Rubber Company, Halmstads Gummifabrik, Woodsafe, ZilenZio, "
                 "Willa Nordic, Nydala Trävaru.\n>\n"
                 "> - **Controllo di terzi** — la società è dentro un gruppo industriale o un fondo "
                 "che decide altrove, spesso all'estero. **Qui il lead perde valore**, ed è il caso "
                 "che il progetto aveva già trattato rimuovendo Lavazza Kaffee e Segafredo Zanetti "
                 "Austria. Rientrano qui Tjørnehøj Mølle (DLG), Skjern Paper (Sonoco, USA), "
                 "Papierfabriek Schut (Exacompta Clairefontaine, FR), Bangma (Stora Enso), "
                 "Timberman (Volati AB), JKE Design (Ballingslöv/Stena Adactum), Bøjsø (Inwido), "
                 "VårgårdaHus (HusCompagniet, DK), Bording (F E Bording, DK), Benecke Coffee "
                 "(Sucafina, CH), Corné Port-Royal (Neuhaus), Sas NV (Nimbus), H. Heitz (INDUS "
                 "Holding), CWP (Auroora Yhtiöt).\n>\n"
                 "> **Il sottoinsieme più grave** non è nessuno dei due, ma i record che "
                 "**affermano un'indipendenza che non c'è**: Tärnsjö Garveri («principale conceria "
                 "indipendente»), Papierfabriek Schut («cartiera indipendente dal 1618»), Billes "
                 "Tryckeri («tipografia indipendente»), Skjern Paper («indipendente danese dal "
                 "2005»), Allbox («azienda familiare»). Lì il foglio non tace un dato: ne afferma "
                 "uno falso.\n")
        o.append("| Foglio | Azienda | Stato nel foglio | Rilievo |")
        o.append("|---|---|---|---|")
        for r in sorted(gr, key=lambda x: (str(x.get("foglio")), str(x.get("denominazione")))):
            k = (r.get("foglio"), r.get("denominazione"))
            rec = recs.get(k)
            if rec is None:
                import unicodedata
                def fold(x):
                    x = (x or "").lower().replace("ø","o").replace("æ","ae").replace("å","aa")
                    x = x.replace("oe","o").replace("ae","a").replace("aa","a")
                    x = unicodedata.normalize("NFKD", x)
                    x = "".join(ch for ch in x if not unicodedata.combining(ch))
                    return re.sub(r"[^a-z0-9]", "", x)
                target = fold(r.get("denominazione"))
                for (sh, dn), v in recs.items():
                    if sh != r.get("foglio"): continue
                    f = fold(dn)
                    if f.startswith(target[:16]) or target.startswith(f[:16]):
                        rec = v; break
            stato = "— (record non risolto)" if rec is None else (
                "**dichiarato**" if GRP.search(rec.get("dimensione","")) else "**NON dichiarato**")
            o.append(f"| {esc(r.get('foglio'))} | {esc(r.get('denominazione'))[:40]} | {stato} | "
                     f"{esc(r.get('problema'))[:200]} |")

    # ---------- casi alta ----------
    alte = [r for r in R if r["gravita"] == "alta"]
    o.append(f"\n---\n\n## 4. Casi di gravità ALTA ({len(alte)})\n")
    o.append("_Dato falso, azienda non contattabile, azienda cessata/fallita/acquisita, "
             "oppure fuori dal perimetro dell'Allegato I EUDR._\n")
    if not alte: o.append("_Nessuno._")
    for sn in SHEETS + altri:
        rs = [r for r in alte if r.get("foglio") == sn]
        if not rs: continue
        o.append(f"\n### {sn} ({len(rs)})\n")
        for r in sorted(rs, key=lambda x: str(x.get("denominazione",""))):
            o.append(f"#### {esc(r.get('denominazione'))} — campo `{esc(r.get('campo'))}`\n")
            o.append(f"{esc(r.get('problema'))}\n")
            o.append(f"**Evidenza:** {esc(r.get('evidenza'))}\n")
            cp = esc(r.get("correzione_proposta"))
            o.append(f"**Correzione proposta:** {cp if cp else '— nessun valore certo: rilievo lasciato aperto'}\n")

    # ---------- media / bassa ----------
    for g, tit, nota in (
        ("media","5. Casi di gravità MEDIA","_Dato dubbio o obsoleto: da rinfrescare prima del contatto, non necessariamente errato._"),
        ("bassa","6. Casi di gravità BASSA","_Refusi formali e incoerenze di stile._")):
        rs_all = [r for r in R if r["gravita"] == g]
        o.append(f"\n---\n\n## {tit} ({len(rs_all)})\n")
        o.append(nota + "\n")
        if not rs_all: o.append("_Nessuno._"); continue
        for sn in SHEETS + altri:
            rs = [r for r in rs_all if r.get("foglio") == sn]
            if not rs: continue
            o.append(f"\n### {sn} ({len(rs)})\n")
            o.append("| Azienda | Campo | Problema | Evidenza | Correzione proposta |")
            o.append("|---|---|---|---|---|")
            for r in sorted(rs, key=lambda x: str(x.get("denominazione",""))):
                o.append(f"| {esc(r.get('denominazione'))[:44]} | {esc(r.get('campo'))} | "
                         f"{esc(r.get('problema'))[:250]} | {esc(r.get('evidenza'))[:150]} | "
                         f"{esc(r.get('correzione_proposta'))[:120]} |")

    md = "\n".join(o) + "\n"
    open(os.path.join(HERE, "_corpo_report.md"), "w", encoding="utf-8").write(md)
    print(f"corpo report: {len(R)} rilievi, {len(files)} blocchi, copertura {tot_ver}/{NTOT}")
    return R, files, tot_ver

if __name__ == "__main__":
    main()

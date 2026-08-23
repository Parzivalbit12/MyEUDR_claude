#!/usr/bin/env python3
"""Fase A - controlli deterministici (offline) su workbook MyEUDR + JSON di build."""
import glob, html, json, os, re, sys, unicodedata
from collections import defaultdict
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # _myeudr_build
REPO = os.path.dirname(ROOT)
XLSX = os.path.join(REPO, "MyEUDR_Lead_Mapping.xlsx")
KEYS = ["denominazione","filiera","dimensione","referente","ruolo","linkedin","email","sito","sede","fonte"]

MACRO = ["Legno/Arredo","Legno/Segheria","Carta/Packaging","Caffè","Cacao/Cioccolato",
         "Gomma","Mangimi/Soia","Bovini/Carne","Pelle/Concia","Olio di palma"]

LEGAL = {
 "Italia":   [r"\bS\.?r\.?l\.?\b", r"\bS\.?p\.?A\.?\b", r"\bS\.?n\.?c\.?\b", r"\bS\.?a\.?s\.?\b", r"\bS\.?o\.?c\.?\b", r"\bcoop", r"\bS\.?S\.?\b"],
 "Germania": [r"\bGmbH\b", r"\bAG\b", r"\bKG\b", r"\bOHG\b", r"\be\.?G\.?\b", r"\bSE\b", r"\bKGaA\b", r"\bmbH\b"],
 "Austria":  [r"\bGmbH\b", r"\bAG\b", r"\bKG\b", r"\bOG\b", r"\be\.?U\.?\b", r"\bGesmbH\b", r"\bmbH\b", r"\beG\b", r"Gesellschaft m\.? ?b\.? ?H", r"\bGes\.?m\.?b\.?H"],
 "Danimarca":[r"\bA/S\b", r"\bApS\b", r"\bAps\b", r"\bA\.?M\.?B\.?A", r"\bI/S\b", r"\bK/S\b", r"\bP/S\b", r"\bAmba\b", r"\ba\.m\.b\.a\.?"],
 "Svezia":   [r"\bAB\b", r"\bHB\b", r"\bKB\b", r"\bek\.? för", r"Aktiebolag", r"\bAB$", r"\bab\b"],
 "Olanda":   [r"\bB\.?V\.?\b", r"\bN\.?V\.?\b", r"\bV\.?O\.?F\.?\b", r"\bC\.?V\.?\b"],
 "Belgio":   [r"\bNV\b", r"\bBV\b", r"\bSA\b", r"\bSRL\b", r"\bSPRL\b", r"\bBVBA\b", r"\bCVBA\b", r"\bCV\b", r"\bComm\.? ?VA?\b", r"\bSCRL\b"],
 "Finlandia":[r"\bOy\b", r"\bOyj\b", r"\bAb\b", r"\bky\b", r"\bosk\b"],
}
COUNTRY_TLD = {"Italia":".it","Germania":".de","Austria":".at","Danimarca":".dk","Svezia":".se",
               "Olanda":".nl","Belgio":".be","Finlandia":".fi"}
REG_RE = re.compile(r"\b(CVR|KVK|org\.?\s?nr|orgnr|organisationsnummer|Firmenbuch|FN\s?\d|BTW|VAT|P\.?\s?IVA|ondernemingsnummer|BE\s?0\d)", re.I)
ENT_RE = re.compile(r"&(amp|gt|lt|quot|apos|nbsp|#\d+);", re.I)

def norm_name(s):
    s = re.sub(r"\(.*?\)", "", s or "")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    for w in ["a/s","aps","gmbh & co. kg","gmbh","ag","kg","ohg","b.v.","bv","n.v.","nv","ab","oy","oyj",
              "s.r.l.","srl","s.p.a.","spa","sa","srl","sprl","bvba","cvba","holding","group","gruppe",
              "aktiebolag","company","co.","& co","international","danmark","sverige","nederland","austria",
              "deutschland","italia","belgium","belgie"]:
        s = s.replace(w, " ")
    return re.sub(r"[^a-z0-9]", "", s)

def dom(u):
    if not u: return ""
    u = u.strip().lower()
    if u in ("n.d.", "n.d", "nd", "-", "n/a"): return ""
    u = re.sub(r"^https?://", "", u)
    u = re.sub(r"^www\.", "", u)
    u = u.split("/")[0].split("?")[0].split(":")[0]
    return u

def base_dom(d):
    """dominio registrabile approssimato (ultimi 2 label, 3 per co.uk-like)"""
    p = d.split(".")
    if len(p) >= 3 and p[-2] in ("co","com","org","net","ac","gov"):
        return ".".join(p[-3:])
    return ".".join(p[-2:]) if len(p) >= 2 else d

# ---------------- caricamento ----------------
records = []   # dict con _src, _sheet
wb = load_workbook(XLSX)
for sn in wb.sheetnames:
    ws = wb[sn]
    for i, r in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not r or not r[0]: continue
        rec = {k: (v or "") if isinstance(v, str) else ("" if v is None else str(v)) for k, v in zip(KEYS, r)}
        rec["_sheet"] = sn; rec["_src"] = f"xlsx:{sn}!r{i}"; rec["_row"] = i
        records.append(rec)

jsons = []
PREF2SHEET = {"dk":"Danimarca","se":"Svezia","nl":"Olanda","be":"Belgio","at":"Austria"}
for fp in sorted(glob.glob(os.path.join(ROOT, "*_*.json"))):
    bn = os.path.basename(fp)
    pref = bn.split("_")[0]
    if pref not in PREF2SHEET: continue
    with open(fp, encoding="utf-8") as f:
        data = json.load(f)
    for j, r in enumerate(data):
        rec = {k: (r.get(k) or "") for k in KEYS}
        rec["_sheet"] = PREF2SHEET[pref]; rec["_src"] = f"{bn}[{j}]"; rec["_file"] = bn
        jsons.append(rec)

findings = defaultdict(list)   # sezione -> lista righe

def add(sec, sheet, den, msg, src):
    findings[sec].append((sheet, den, msg, src))

# ---------------- 1. DUPLICATI ----------------
by_key = defaultdict(list)
for r in records:
    by_key[norm_name(r["denominazione"])].append(r)
for k, rs in sorted(by_key.items()):
    if len(rs) > 1:
        add("1_duplicati_xlsx", rs[0]["_sheet"], rs[0]["denominazione"],
            "duplicato su " + " | ".join(f'{x["_sheet"]}: "{x["denominazione"]}"' for x in rs), "")

by_key_j = defaultdict(list)
for r in jsons:
    by_key_j[norm_name(r["denominazione"])].append(r)
for k, rs in sorted(by_key_j.items()):
    if len(rs) > 1:
        files = {x["_file"] for x in rs}
        add("1b_duplicati_json", rs[0]["_sheet"], rs[0]["denominazione"],
            f"{len(rs)} occorrenze in " + ", ".join(sorted(files)), "")

# domini identici su denominazioni diverse (possibile stessa azienda / gruppo)
by_dom = defaultdict(set)
for r in records:
    d = base_dom(dom(r["sito"]))
    if d and "." in d:
        by_dom[d].add((r["_sheet"], r["denominazione"]))
for d, s in sorted(by_dom.items()):
    if len(s) > 1:
        add("1c_stesso_dominio", "/".join(sorted({x[0] for x in s})), d,
            "stesso sito per: " + " | ".join(f"{a}: {b}" for a, b in sorted(s)), "")

# ---------------- 2. EMAIL vs DOMINIO SITO ----------------
FREE = {"gmail.com","hotmail.com","outlook.com","yahoo.com","live.com","libero.it","pec.it",
        "legalmail.it","aol.com","icloud.com","gmx.de","gmx.at","gmx.net","web.de","telenet.be",
        "skynet.be","planet.nl","ziggo.nl","kpnmail.nl","mail.dk","tiscali.it","virgilio.it",
        "alice.it","t-online.de","chello.at","aon.at","a1.net","hotmail.it","pec.legalmail.it"}
for r in records:
    em = r["email"].strip()
    if not em or em == "n.d." or "@" not in em: continue
    ed = base_dom(em.split("@")[-1].strip().lower())
    sd = base_dom(dom(r["sito"]))
    if not sd: 
        add("2b_email_senza_sito", r["_sheet"], r["denominazione"], f"email {em} ma campo sito vuoto", r["_src"])
        continue
    if ed != sd:
        es, ss = ed.split(".")[0], sd.split(".")[0]
        if ed in FREE:
            tag, sec = "FREEMAIL/PEC", "2c_email_freemail"
        elif es == ss:
            tag, sec = "STESSO NOME, TLD DIVERSO", "2d_email_tld_diverso"
        elif es in ss or ss in es or len(set(es) & set(ss)) >= max(3, min(len(es), len(ss)) - 3) and (es[:5] == ss[:5]):
            tag, sec = "STEM AFFINE", "2d_email_tld_diverso"
        else:
            tag, sec = "STEM DIVERSO — DA CONFERMARE", "2_email_dominio_diverso"
        add(sec, r["_sheet"], r["denominazione"],
            f"[{tag}] email @{ed} vs sito {sd}  (email: {em} | sito: {r['sito']})", r["_src"])

# ---------------- 3. EMAIL / LINKEDIN RIPETUTI ----------------
for field, sec in (("email","3_email_ripetute"), ("linkedin","3b_linkedin_ripetuti")):
    m = defaultdict(set)
    for r in records:
        v = r[field].strip().lower()
        if not v or v == "n.d.": continue
        m[v].add((r["_sheet"], r["denominazione"]))
    for v, s in sorted(m.items()):
        if len(s) > 1:
            add(sec, "/".join(sorted({x[0] for x in s})), v,
                "condiviso da: " + " | ".join(f"{a}: {b}" for a, b in sorted(s)), "")

# ---------------- 4. URL MALFORMATI ----------------
for r in records:
    li = r["linkedin"].strip()
    if li and li.lower() != "n.d." and "linkedin.com" not in li.lower():
        add("4_url_malformati", r["_sheet"], r["denominazione"], f"LinkedIn non linkedin.com: {li}", r["_src"])
    if li and li.lower() != "n.d." and not li.lower().startswith("http"):
        add("4_url_malformati", r["_sheet"], r["denominazione"], f"LinkedIn senza http: {li}", r["_src"])
    st = r["sito"].strip()
    if st and st.lower() != "n.d." and not st.lower().startswith("http"):
        add("4_url_malformati", r["_sheet"], r["denominazione"], f"Sito senza http: {st}", r["_src"])
    em = r["email"].strip()
    if em and em.lower() != "n.d.":
        if not re.fullmatch(r"[^@\s,;]+@[^@\s,;]+\.[a-z]{2,}", em, re.I):
            add("4b_email_malformate", r["_sheet"], r["denominazione"], f"email non conforme: {em}", r["_src"])
    fo = r["fonte"].strip()
    if fo and not fo.lower().startswith("http"):
        add("7_fonte", r["_sheet"], r["denominazione"], f"Fonte non URL: {fo[:90]}", r["_src"])
    if not fo:
        add("7_fonte", r["_sheet"], r["denominazione"], "Fonte VUOTA", r["_src"])
    if not dom(r["sito"]):
        add("7c_sito_mancante", r["_sheet"], r["denominazione"], f"Sito = '{r['sito']}'", r["_src"])
    if not r["dimensione"].strip() or r["dimensione"].strip().lower() == "n.d.":
        add("7b_dimensione_vuota", r["_sheet"], r["denominazione"], f"Dimensione = '{r['dimensione']}'", r["_src"])

# ---------------- 5. ENTITA' HTML ----------------
for pool, tag in ((records, "xlsx"), (jsons, "json")):
    for r in pool:
        for k in KEYS:
            m = ENT_RE.search(r[k] or "")
            if m:
                add("5_entita_html", r["_sheet"], r["denominazione"],
                    f"[{tag}] campo {k}: {m.group(0)} -> ...{(r[k])[max(0,m.start()-30):m.start()+30]}...", r["_src"])

# ---------------- 6. TASSONOMIA FILIERA ----------------
macro_counts = defaultdict(int)
for r in records:
    f = (r["filiera"] or "").strip()
    if not f:
        add("6_filiera", r["_sheet"], r["denominazione"], "Filiera VUOTA", r["_src"]); continue
    parts = re.split(r"\s+[—–-]\s+", f, maxsplit=1)
    macro = parts[0].strip()
    macro_counts[macro] += 1
    if macro not in MACRO:
        add("6_filiera", r["_sheet"], r["denominazione"], f"macro fuori tassonomia: «{macro}»  (valore intero: «{f}»)", r["_src"])
    if "—" not in f and len(parts) > 1:
        add("6b_filiera_separatore", r["_sheet"], r["denominazione"], f"separatore non em-dash: «{f}»", r["_src"])

# ---------------- 8. DIMENSIONE: numeri fuori forbice ----------------
RATE = {"DKK":7.46, "SEK":11.3, "NOK":11.5, "EUR":1.0}
def euros(r):
    """estrae candidati fatturato in M€ dal campo dimensione"""
    t = r["dimensione"]
    out = []
    # pattern:  123,4 M DKK / 12,5 M€ / 1,2 mld DKK / 45 MSEK / 10 820 KSEK / EUR 12 Mio
    for m in re.finditer(r"(\d[\d\.\s\u00a0]*(?:,\d+)?)\s*(mld|miliardi|mia\.?|mrd)?\s*"
                         r"(M€|MEUR|M EUR|Mio\.? ?€|Mio\.? ?EUR|M\s?DKK|MDKK|M\s?SEK|MSEK|KSEK|TSEK|TDKK|KDKK|"
                         r"mio\.? ?€|milioni di euro|M\$|€|DKK|SEK|EUR)\b", t, re.I):
        num = m.group(1).replace(".", "").replace(" ", "").replace("\u00a0", "").replace(",", ".")
        try: v = float(num)
        except ValueError: continue
        unit = (m.group(3) or "").upper().replace(" ", "").replace(".", "")
        mld = bool(m.group(2))
        cur = "EUR"
        if "DKK" in unit: cur = "DKK"
        elif "SEK" in unit: cur = "SEK"
        scale = 1.0                    # in milioni di valuta
        if unit.startswith("K") or unit.startswith("T"):   # KSEK/TSEK/KDKK migliaia
            scale = 0.001
        elif unit in ("€","DKK","SEK","EUR"):              # unità piene
            scale = 1e-6
        if mld: scale = 1000.0
        out.append((v * scale / RATE[cur], m.group(0)))
    return out

SIZE_FLAG = re.compile(r"sopra soglia|fuori soglia|sotto (i |la )?(soglia|target|5)|oltre (la )?soglia|"
                       r"da verificare|stima|non pubblicat|bruttofortjeneste|totale di bilancio|"
                       r"segnalat|sopra ?fascia|sopra la fascia|fuori fascia|sotto fascia|oltre fascia|"
                       r"gruppo|controllata|capogruppo|volumi|fm/anno|indipendente ma|micro|patrimonio", re.I)
for r in records:
    cands = euros(r)
    if not cands: continue
    vals = [v for v, _ in cands]
    inband = [v for v in vals if 5 <= v <= 40]
    if inband: continue
    if SIZE_FLAG.search(r["dimensione"]): continue
    add("8_dimensione_fuori_forbice", r["_sheet"], r["denominazione"],
        "valori stimati M€: " + ", ".join(f"{v:.2f} («{s.strip()}»)" for v, s in cands) +
        f" | campo: {r['dimensione'][:150]}", r["_src"])

# ---------------- 9. DENOMINAZIONE ----------------
for r in records:
    d = r["denominazione"]
    m = REG_RE.search(d)
    if m:
        add("9_denominazione", r["_sheet"], d, f"numero/sigla di registro nel nome: «{m.group(0)}»", r["_src"])
    if "  " in d:
        add("9_denominazione", r["_sheet"], d, "doppio spazio nel nome", r["_src"])
    if d != d.strip():
        add("9_denominazione", r["_sheet"], d, "spazi iniziali/finali", r["_src"])
    if re.search(r"\d{6,}", d):
        add("9_denominazione", r["_sheet"], d, "sequenza numerica lunga nel nome", r["_src"])
    # forma giuridica coerente col paese
    sheet = r["_sheet"]
    pats = LEGAL.get(sheet, [])
    if pats and not any(re.search(p, d) for p in pats):
        # verifica se ha una forma giuridica di un ALTRO paese
        other = [c for c, ps in LEGAL.items() if c != sheet and any(re.search(p, d) for p in ps)]
        if other:
            add("9b_forma_giuridica", sheet, d, f"forma giuridica tipica di {', '.join(other)} in foglio {sheet}", r["_src"])
        else:
            add("9c_forma_giuridica_assente", sheet, d, "nessuna forma giuridica riconoscibile", r["_src"])

# ---------------- 10. TLD sito/email incoerente col paese ----------------
GENERIC = {".com",".eu",".net",".org",".coffee",".shop",".group",".info",".biz",".online",".store",".io",".co"}
for r in records:
    sd = dom(r["sito"])
    if not sd: continue
    tld = "." + sd.split(".")[-1]
    exp = COUNTRY_TLD.get(r["_sheet"])
    if exp and not sd.endswith(exp) and tld not in GENERIC:
        add("10_tld_estero", r["_sheet"], r["denominazione"], f"sito {r['sito']} — TLD {tld} estraneo a {r['_sheet']}", r["_src"])

# ---------------- 11. divergenze JSON vs XLSX ----------------
xl_by = {}
for r in records:
    xl_by[(r["_sheet"], norm_name(r["denominazione"]))] = r
for r in jsons:
    k = (r["_sheet"], norm_name(r["denominazione"]))
    x = xl_by.get(k)
    if not x:
        add("11b_json_non_nel_foglio", r["_sheet"], r["denominazione"], f"presente in {r['_file']} ma NON nel foglio", r["_src"])
        continue
    for f in ("email","sito","linkedin","referente","filiera"):
        a = html.unescape(str(r[f] or "")).strip()
        b = str(x[f] or "").strip()
        if a and b and a != b and a != "n.d." and b != "n.d.":
            add("11_divergenze_json_xlsx", r["_sheet"], r["denominazione"],
                f"campo {f}: JSON «{a[:70]}» ≠ XLSX «{b[:70]}»", r["_src"])

# ---------------- 12. campi vuoti (statistica) ----------------
stats = defaultdict(lambda: defaultdict(int))
for r in records:
    stats[r["_sheet"]]["_n"] += 1
    for f in ("email","referente","ruolo","linkedin","sito","sede","fonte","dimensione"):
        v = r[f].strip()
        if v and v.lower() != "n.d.":
            stats[r["_sheet"]][f] += 1

# ---------------- OUTPUT ----------------
TITLES = {
 "1_duplicati_xlsx":"1 · Duplicati fra fogli del workbook",
 "1b_duplicati_json":"1b · Duplicati fra i JSON di build",
 "1c_stesso_dominio":"1c · Denominazioni diverse con lo stesso sito web",
 "2_email_dominio_diverso":"2 · Email con dominio diverso dal sito (sospette di deduzione)",
 "2b_email_senza_sito":"2b · Email presente ma sito assente (non verificabile per dominio)",
 "2c_email_freemail":"2c · Email su dominio freemail/PEC (accettabile ma non aziendale)",
 "2d_email_tld_diverso":"2d · Email su dominio affine al sito (TLD/variante) — rischio basso",
 "3_email_ripetute":"3 · Stessa email su aziende diverse",
 "3b_linkedin_ripetuti":"3b · Stesso LinkedIn su aziende diverse",
 "4_url_malformati":"4 · URL malformati (LinkedIn / sito)",
 "4b_email_malformate":"4b · Email sintatticamente non conformi",
 "5_entita_html":"5 · Entità HTML residue",
 "6_filiera":"6 · Tassonomia Filiera fuori elenco",
 "6b_filiera_separatore":"6b · Separatore filiera non em-dash",
 "7_fonte":"7 · Fonte vuota o non URL",
 "7b_dimensione_vuota":"7b · Dimensione vuota o n.d.",
 "7c_sito_mancante":"7c · Sito web mancante",
 "8_dimensione_fuori_forbice":"8 · Dimensione fuori forbice 5–40 M€ senza segnalazione esplicita",
 "9_denominazione":"9 · Denominazione: registri, spazi, numeri",
 "9b_forma_giuridica":"9b · Forma giuridica incoerente col paese del foglio",
 "9c_forma_giuridica_assente":"9c · Nessuna forma giuridica nel nome",
 "10_tld_estero":"10 · TLD del sito estraneo al paese del foglio",
 "11_divergenze_json_xlsx":"11 · Divergenze fra JSON di build e foglio Excel",
 "11b_json_non_nel_foglio":"11b · Record presente nei JSON ma assente dal foglio",
}
ORDER_SEC = list(TITLES.keys())

out = []
out.append("# Fase A — Controlli deterministici automatici\n")
out.append("_Generato da `_myeudr_build/verifica/controlli_automatici.py` · nessun accesso di rete._\n")
out.append(f"**Perimetro:** {len(records)} record nel workbook ({len(wb.sheetnames)} fogli) + "
           f"{len(jsons)} record nei {len(set(r['_file'] for r in jsons))} file JSON di build.\n")

out.append("\n## Copertura campi per foglio\n")
out.append("| Foglio | Record | Email | Referente | Ruolo | LinkedIn | Sito | Sede | Fonte | Dimensione |")
out.append("|---|--:|--:|--:|--:|--:|--:|--:|--:|--:|")
for sn in wb.sheetnames:
    s = stats[sn]; n = s["_n"]
    out.append(f"| {sn} | {n} | {s['email']} | {s['referente']} | {s['ruolo']} | {s['linkedin']} | "
               f"{s['sito']} | {s['sede']} | {s['fonte']} | {s['dimensione']} |")

out.append("\n## Riepilogo rilievi\n")
out.append("| # | Controllo | Rilievi |")
out.append("|---|---|--:|")
for sec in ORDER_SEC:
    out.append(f"| {sec.split('_')[0]} | {TITLES[sec]} | {len(findings.get(sec, []))} |")
out.append(f"\n**Totale rilievi automatici: {sum(len(v) for v in findings.values())}**\n")

out.append("\n## Macro-filiere osservate\n")
out.append("| Macro | Occorrenze | In tassonomia |")
out.append("|---|--:|:--:|")
for m, c in sorted(macro_counts.items(), key=lambda x: -x[1]):
    out.append(f"| {m} | {c} | {'✅' if m in MACRO else '❌'} |")

for sec in ORDER_SEC:
    rows = findings.get(sec, [])
    out.append(f"\n---\n\n## {TITLES[sec]}  ({len(rows)})\n")
    if not rows:
        out.append("_Nessun rilievo._"); continue
    out.append("| Foglio | Denominazione | Rilievo | Origine |")
    out.append("|---|---|---|---|")
    for sheet, den, msg, src in rows:
        esc = lambda s: str(s).replace("|", "\\|").replace("\n", " ")
        out.append(f"| {esc(sheet)} | {esc(den)[:60]} | {esc(msg)[:400]} | {esc(src)} |")

md = "\n".join(out) + "\n"
op = os.path.join(ROOT, "verifica", "00_controlli_automatici.md")
open(op, "w", encoding="utf-8").write(md)
print("scritto", op, len(md), "byte")
for sec in ORDER_SEC:
    print(f"  {len(findings.get(sec,[])):4d}  {TITLES[sec]}")

# dump JSON dei record per gli agenti di Fase B
dump = os.path.join(ROOT, "verifica", "_records.json")
json.dump([{k: r[k] for k in KEYS + ["_sheet","_row"]} for r in records],
          open(dump, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("scritto", dump)

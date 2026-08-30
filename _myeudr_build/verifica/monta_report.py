#!/usr/bin/env python3
"""Monta REPORT_VERIFICA.md: intestazione + sintesi Fase A + correzioni applicate + corpo Fase B."""
import glob, json, os, re, subprocess, sys

HERE = os.path.dirname(os.path.abspath(__file__))
subprocess.run([sys.executable, os.path.join(HERE, "genera_report.py")], check=True)
corpo = open(os.path.join(HERE, "_corpo_report.md"), encoding="utf-8").read()

# --- sintesi Fase A dal report automatico ---
auto = open(os.path.join(HERE, "00_controlli_automatici.md"), encoding="utf-8").read()
riepilogo = re.search(r"## Riepilogo rilievi\n(.*?)\n\*\*Totale", auto, re.S)
tab_A = riepilogo.group(1).strip() if riepilogo else ""
tot_A = re.search(r"\*\*Totale rilievi automatici: (\d+)\*\*", auto)

# --- correzioni applicate ---
corr = []
REF_FILE = os.path.join(HERE, "correzioni_referenti.json")
for fp in sorted(glob.glob(os.path.join(HERE, "correzioni_*.json"))):
    if os.path.abspath(fp) == os.path.abspath(REF_FILE): continue   # schema diverso, sezione dedicata
    corr += json.load(open(fp, encoding="utf-8"))
try:
    refs = json.load(open(REF_FILE, encoding="utf-8"))
except Exception:
    refs = []

def esc(s): return str(s or "").replace("|", "\\|").replace("\n", " ").strip()

h = []
h.append("# REPORT DI VERIFICA — MyEUDR Lead Mapping\n")
from openpyxl import load_workbook as _lw
_wb = _lw(os.path.join(os.path.dirname(os.path.dirname(HERE)), "MyEUDR_Lead_Mapping.xlsx"), read_only=True)
NTOT = sum(sum(1 for r in _wb[sn].iter_rows(min_row=3, values_only=True) if r and r[0]) for sn in _wb.sheetnames)
h.append(f"> Controllo qualità **record per record** del censimento lead (**{NTOT} aziende, 8 fogli**), "
         "alla ricerca di refusi, attribuzioni errate e ogni altro errore introdotto durante la "
         "raccolta. Non è una ricerca di nuove aziende.\n")
h.append("\n## Come leggere questo report\n")
h.append("La verifica si è svolta in due fasi, con budget e coperture diverse:\n")
h.append("| Fase | Metodo | Copertura |")
h.append("|---|---|---|")
h.append(f"| **A — controlli deterministici** | 26 controlli automatici offline su tutti i JSON di "
         f"build e sul workbook | **100%** dei {NTOT} record |")
h.append("| **B — riscontro sul web** | agenti di verifica, blocchi di 15-20 aziende, 2-3 ricerche "
         "per record, ogni rilievo con URL o citazione | vedi §1 |")
h.append("\nDocumenti di dettaglio:\n")
h.append("- [`00_controlli_automatici.md`](00_controlli_automatici.md) — esito completo della Fase A")
h.append("- [`01_analisi_dimensione.md`](01_analisi_dimensione.md) — analisi del campo `Dimensione` "
         "(tipo di dato dichiarato, anno, obsolescenza)")
h.append("- `<foglio>_<blocco>.json` — rilievi grezzi di ciascun agente")
h.append("- `00_punti_noti.json` — verifica dei 13 punti lasciati aperti dalla raccolta\n")

h.append("\n## Vincolo d'ambiente che ha condizionato la verifica\n")
h.append("Il proxy di egress **nega per policy** (403 al CONNECT) tutti i domini esterni: registri "
         "societari e siti aziendali non sono raggiungibili né con WebFetch né con curl. "
         "**L'unico canale disponibile è WebSearch**, i cui frammenti contengono spesso — ma non "
         "sempre — il dato cercato. Di qui una regola applicata da tutti gli agenti: quando dopo "
         "2-3 ricerche il dato non emerge, il rilievo è marcato `DA CONFERMARE` e il campo "
         "**non viene toccato**. Un rilievo aperto è preferibile a una correzione inventata.\n")

h.append("\n---\n")
h.append("\n## 0. Esito della Fase A — controlli deterministici\n")
h.append("Il censimento è risultato **pulito su tutti i controlli di integrità**: nessun duplicato "
         "(né fra fogli né fra i JSON di build), nessuna entità HTML residua, nessun URL LinkedIn "
         "o sito malformato, nessuna email sintatticamente errata, nessun numero di registro "
         "rimasto nelle denominazioni, nessun campo `Fonte` vuoto o non-URL, e **nessuna divergenza "
         "fra i JSON di build e i fogli Excel** — segno che la pipeline `add_country.py` è "
         "rimasta coerente.\n")
h.append("I rilievi si concentrano invece su qualità e coerenza redazionale:\n")
h.append("> La tabella riflette lo stato **dopo** le correzioni applicate (§0-bis): per questo il "
         "controllo 6 sulla tassonomia è ora a zero, mentre prima dell'intervento segnalava "
         "**23 valori fuori elenco**.\n")
h.append(tab_A + "\n")
if tot_A:
    h.append(f"\n**Totale rilievi automatici: {tot_A.group(1)}.**\n")
h.append("\n> **Nota sul controllo 2 (email dedotte).** Il controllo grezzo segnalava 86 email con "
         "dominio diverso dal sito. Separando i casi innocui — stesso nome con TLD diverso "
         "(`azienda.de` vs `azienda.com`) e caselle freemail/PEC, entrambi legittimi — restano "
         "**24 casi con stem realmente diverso**, che sono quelli da confermare a fonte.\n")

if corr:
    h.append("\n---\n")
    h.append(f"\n## 0-bis. Correzioni già applicate al workbook ({len(corr)})\n")
    h.append("Applicate **solo le correzioni certe**, secondo il mandato: refusi formali, entità "
             "HTML, forme giuridiche, filiere fuori Allegato I, aziende cessate. Tutto il resto "
             "resta come rilievo aperto in questo report.\n")
    h.append("Ogni correzione di campo è stata applicata con un **controllo di guardia**: lo script verifica "
             "che il valore attuale del campo coincida esattamente con quello atteso, altrimenti "
             "salta la correzione, così lo script è rieseguibile senza rischi. Dopo l'applicazione le "
             "righe sono **740** (due rimozioni motivate, vedi sotto) e l'ordine dei fogli è "
             "ripristinato (Italia, Germania, Finlandia, Danimarca, Svezia, Olanda, Belgio, Austria).\n")
    rim  = [c for c in corr if c["a"] is None]
    perf = [c for c in corr if c["a"] is not None and c["campo"] == "filiera"]
    altre = [c for c in corr if c["a"] is not None and c["campo"] != "filiera"]
    if rim:
        h.append(f"\n### Record rimossi dal censimento ({len(rim)})\n")
        h.append("Sono le uniche righe **tolte** dai fogli. Ciascuna rientra in una categoria che "
                 "il mandato autorizza a correggere — filiere fuori Allegato I, aziende cessate — "
                 "e in ogni caso il progetto aveva già applicato lo stesso criterio a un caso "
                 "analogo, che viene citato nella motivazione.\n")
        for c in rim:
            h.append(f"**{esc(c['denominazione'])}** — foglio {esc(c['foglio'])}  ")
            h.append(f"{esc(c['motivo'])}\n")
        h.append(f"Il totale del censimento passa quindi da **742 a {NTOT} aziende** "
                 "(Belgio 95→94, Olanda 100→99, Austria 93→92).\n")
    h.append(f"\n### Tassonomia `Filiera` ({len(perf)})\n")
    h.append("Il foglio **Finlandia** conteneva varianti storiche della tassonomia "
             "(`Legno/Compensato-Prodotti`, `Legno/Segheria-Piallatura`, `Legno/CLT`, "
             "`Legno/Commercio-export sahatavara`, `Legno/Piallatura`) mai allineate agli altri "
             "fogli. La riconduzione **non è arbitraria**: gli altri sette fogli sono unanimi nel "
             "classificare compensato, impiallacciature, finestre/porte, parquet, glulam e CLT "
             "sotto `Legno/Arredo — <dettaglio>`, e segheria/piallatura sotto `Legno/Segheria`.\n")
    h.append("| Foglio | Azienda | Da | A |")
    h.append("|---|---|---|---|")
    for c in perf:
        h.append(f"| {esc(c['foglio'])} | {esc(c['denominazione'])[:38]} | {esc(c['da'])[:52]} | {esc(c['a'])[:52]} |")
    if altre:
        h.append(f"\n### Refusi formali ({len(altre)})\n")
        h.append("| Foglio | Azienda | Campo | Correzione | Motivo |")
        h.append("|---|---|---|---|---|")
        for c in altre:
            da, a = esc(c["da"]), esc(c["a"])
            if len(da) > 70:   # mostra solo il frammento che cambia
                import difflib
                sm = difflib.SequenceMatcher(None, da, a)
                fr = [(da[i1:i2], a[j1:j2]) for tag,i1,i2,j1,j2 in sm.get_opcodes() if tag != "equal"]
                testo = " · ".join(f"«{x[:45]}» → «{y[:45]}»" for x, y in fr) or "(riscrittura)"
            else:
                testo = f"«{da}» → «{a}»"
            h.append(f"| {esc(c['foglio'])} | {esc(c['denominazione'])[:32]} | {esc(c['campo'])} | "
                     f"{testo[:170]} | {esc(c['motivo'])[:110]} |")

h.append("\n> ⚠️ **Limite della normalizzazione ortografica del foglio Italia — da tenere presente.** "
         "Le 18 riscritture `Srl`→`S.r.l.` e `SpA`→`S.p.A.` correggono **come è scritta** la forma "
         "giuridica, non **quale** sia: assumono che il foglio l'avesse indovinata. Su **Arko** "
         "l'assunzione era falsa — le fonti camerali danno `ARKO S.R.L.` con la stessa P.IVA "
         "03278760263 — e la normalizzazione l'aveva resa `Arko S.p.A.`, cioè aveva reso più "
         "autorevole una forma sbagliata. Il record è stato corretto dopo la verifica di merito. "
         "Lo stesso può valere per i record dei blocchi Italia **non ancora verificati**: solo il "
         "controllo record per record distingue un refuso di scrittura da una forma giuridica errata.\n>\n"
         "> **Quanto è esteso il problema, per ora.** Tre blocchi Italia verificati (57 record, "
         "tutte le P.IVA confrontate al Registro): **2 forme giuridiche sbagliate**, Fonpelli e "
         "Arko, entrambe nei primi due blocchi. Il terzo blocco ne ha trovate **zero su 19**. "
         "Sembrano quindi casi isolati e non un difetto sistematico del foglio — ma i due blocchi "
         "Italia rimanenti non sono ancora stati controllati.\n")

if refs:
    nref = sum(1 for c in refs if c.get("referente"))
    nalta = sum(1 for c in refs if c.get("gravita") == "alta")
    h.append(f"\n### Referenti e ruoli riversati nei fogli ({len(refs)})\n")
    h.append("Applicati su richiesta del cliente, dopo la prima consegna. Sono i referenti "
             f"**verificati a fonte** in Fase B: {nref} portano un nome, gli altri correggono solo il "
             f"ruolo; {nalta} sono di gravità `alta`.\n")
    h.append("Tre regole hanno governato l'applicazione:\n")
    h.append("1. **Nessun campo viene mai svuotato.** Molte proposte riguardavano solo il *ruolo*: "
             "applicarle alla lettera avrebbe cancellato il nome. Verificato a posteriori: "
             "**0 campi svuotati**.\n")
    h.append("2. **Escluse le proposte con riserva** — 36 su 172 contenevano «DA CONFERMARE», "
             "«da riconfermare», «in alternativa» o un punto interrogativo, più 3 che erano un titolo "
             "e non un nome di persona. Quelle **restano solo qui nel report**.\n")
    h.append("3. **Guardia sul valore attuale**, come per ogni altra correzione: 26 proposte "
             "coincidevano già col foglio e sono state saltate.\n")
    h.append("\nIl guadagno non è tanto nella copertura — i referenti passano da 535 a **575 su 728** "
             "— quanto nella **sostituzione di nomi sbagliati**: predecessori, persone con un ruolo "
             "diverso da quello indicato (il presidente del CdA al posto dell'AD, il direttore "
             "finanziario al posto del CEO) e, in tre casi, **il dirigente di un'altra società**.\n")
    h.append("| Foglio | Azienda | Referente | Ruolo | Gravità |")
    h.append("|---|---|---|---|---|")
    for c in sorted(refs, key=lambda x: (x["foglio"], x["denominazione"])):
        h.append(f"| {esc(c['foglio'])} | {esc(c['denominazione'])[:40]} | "
                 f"{esc(c.get('referente')) or '—'} | {esc(c.get('ruolo')) or '—'} | {c.get('gravita','')} |")

h.append("\n### Il criterio usato per rimuovere, e quello per NON rimuovere\n")
h.append("Gli errori di perimetro non sono tutti uguali, e la differenza decide se una riga esce dal "
         "foglio o resta come rilievo aperto:\n")
h.append("| | Esito | Casi |")
h.append("|---|---|---|")
h.append("| **La commodity è fuori Allegato I** — l'azienda lavora una materia che il regolamento "
         "non copre | **rimossa** | Covera Packaging (vetro e plastica), Vandeputte Oleochemicals "
         "(lino), Helvoet (elastomeri sintetici), Compex (attrezzature zootecniche), Marine Olie "
         "(UCO per biocarburanti), Mejling Landhandel (negozio al dettaglio di generi alimentari) |")
h.append("| **L'azienda non esiste più o non produce più** | **rimossa** | Odense Seglmærkefabrik, "
         "Kaffekompaniet (sciolte per fusione), Weissengruber, Helmut Sachers, Sisuwood (insolvenze), "
         "Bayer Kartonagen, Marandi (produzione cessata) |")
h.append("| **La commodity È in Allegato I e l'unico argomento è la posizione nella filiera** — "
         "l'azienda commercia o rivende invece di immettere per prima | **rilievo aperto** | "
         "Varia-Pack, Hausberger, Pappersgrossisten, Däckteam, Svenska Gummihuset, Cebeco Fourage, "
         "Skovs Korn, Rickl-Mühle, SRC, Kargro Banden |")
h.append("\nLa terza riga è una **scelta deliberata**: stabilire se un commerciante sia «operatore» "
         "ai sensi dell'EUDR è una valutazione giuridica, non un dato di fatto verificabile a fonte. "
         "Quei dieci record restano nel foglio con il rilievo motivato, perché la decisione spetta al "
         "cliente.\n")

h.append("\n### Correzioni deliberatamente NON applicate\n")
h.append("Tre categorie di rilievi formali sono state lasciate aperte nel report invece che "
         "corrette nei fogli. Il motivo è sempre lo stesso: la correzione automatica avrebbe "
         "introdotto un errore nuovo.\n")
h.append("| Rilievo | Record | Perché non è stata applicata |")
h.append("|---|--:|---|")
h.append("| **Maiuscolo integrale nel foglio Danimarca** (controllo 9d) | 52 | Il foglio mescola "
         "51 denominazioni in MAIUSCOLO (stile del registro CVR) e 38 in forma normale. "
         "Un *title case* automatico però **rovinerebbe gli acronimi**: `JKE DESIGN` diventerebbe "
         "`Jke Design`, e lo stesso vale per NPI, MC, KLS, H.C., DHS. Servirebbe una decisione "
         "caso per caso, che non è una correzione certa. |")
h.append("| **Conceria Beschin** e **Conceria Daniela** (foglio Italia) | 2 | Ciascuna corrisponde "
         "a **due entità distinte e omonime** al Registro Imprese, nello stesso comune (una S.n.c. "
         "e una S.r.l.). Aggiungere una forma giuridica significherebbe **scegliere** quale sia "
         "l'operatore EUDR: va accertato prima del contatto. |")
h.append("| **Email da confermare** | 24+ | Le email con dominio diverso dal sito, e quelle non "
         "ritrovate letteralmente in una fonte pubblica, restano **`DA CONFERMARE`**. Il mandato "
         "vieta sia di inventarle sia di cancellarle d'ufficio: il campo non è stato toccato. |")

h.append("\n---\n")
# il corpo comincia col proprio titolo H1: lo tolgo e tengo dalla sezione 1
corpo_body = corpo.split("\n## 1. Copertura della verifica\n", 1)
corpo_finale = "\n## 1. Copertura della verifica\n" + corpo_body[1] if len(corpo_body) > 1 else corpo
h.append("\n# Fase B — verifica sul web, record per record\n")

out = "\n".join(h) + "\n" + corpo_finale
op = os.path.join(HERE, "REPORT_VERIFICA.md")
open(op, "w", encoding="utf-8").write(out)
print(f"scritto {op} — {len(out)} byte, {out.count(chr(10))} righe")

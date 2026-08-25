#!/usr/bin/env python3
"""
Applica al workbook SOLO le correzioni certe registrate nei file correzioni_*.json.

Due percorsi, perché i fogli non hanno tutti la stessa origine:
  · Danimarca/Svezia/Olanda/Belgio/Austria hanno i JSON di build: si corregge il JSON
    e si rigenera il foglio con add_country.py (rigenerazione verificata identica).
  · Italia/Germania/Finlandia esistono SOLO come foglio Excel: si corregge la cella in
    posto, perché una rigenerazione da JSON riordinerebbe le righe già consegnate.

Uso:  python applica_correzioni.py [--dry-run]
"""
import glob, json, os, re, subprocess, sys

HERE = os.path.dirname(os.path.abspath(__file__))
BUILD = os.path.dirname(HERE)
REPO = os.path.dirname(BUILD)
XLSX = os.path.join(REPO, "MyEUDR_Lead_Mapping.xlsx")
ORDINE = ["Italia","Germania","Finlandia","Danimarca","Svezia","Olanda","Belgio","Austria"]
PREFIX = {"Danimarca":"dk","Svezia":"se","Olanda":"nl","Belgio":"be","Austria":"at"}
INPLACE = {"Italia","Germania","Finlandia"}
COL = {"denominazione":1,"filiera":2,"dimensione":3,"referente":4,"ruolo":5,
       "linkedin":6,"email":7,"sito":8,"sede":9,"fonte":10}

DRY = "--dry-run" in sys.argv

def norm(s):
    return re.sub(r"[^a-z0-9]", "", re.sub(r"\(.*?\)", "", (s or "")).lower())

def load_corr():
    C = []
    for fp in sorted(glob.glob(os.path.join(HERE, "correzioni_*.json"))):
        C += json.load(open(fp, encoding="utf-8"))
    return C

def main():
    from openpyxl import load_workbook
    C = load_corr()
    print(f"{len(C)} correzioni da applicare\n")

    # ---- 1. fogli con JSON di build: correggo il JSON ----
    tocchi = {}
    for c in C:
        sh = c["foglio"]
        if sh in INPLACE: continue
        pref = PREFIX[sh]
        done = False
        for fp in sorted(glob.glob(os.path.join(BUILD, f"{pref}_*.json"))):
            data = json.load(open(fp, encoding="utf-8"))
            for r in data:
                if norm(r.get("denominazione")) != norm(c["denominazione"]): continue
                cur = (r.get(c["campo"]) or "")
                if c.get("da") and str(cur).strip() != c["da"].strip():
                    print(f"  ~ {sh}/{c['denominazione']}: valore attuale «{cur[:60]}» "
                          f"≠ atteso «{c['da'][:60]}» — SALTATA")
                    done = True; break
                if c["a"] is None:
                    data.remove(r); print(f"  - {sh}: RIMOSSO record «{c['denominazione']}» ({c['motivo']})")
                else:
                    r[c["campo"]] = c["a"]
                    print(f"  ✔ {sh}/{c['denominazione'][:32]} · {c['campo']}: «{str(cur)[:45]}» → «{c['a'][:45]}»")
                if not DRY:
                    json.dump(data, open(fp,"w",encoding="utf-8"), ensure_ascii=False, indent=1)
                tocchi[sh] = True; done = True; break
            if done: break
        if not done:
            gia = any(norm(r.get("denominazione")) == norm(c["a"])
                      for fp in glob.glob(os.path.join(BUILD, f"{pref}_*.json"))
                      for r in json.load(open(fp, encoding="utf-8"))) if c["campo"] == "denominazione" else False
            print(f"  = {sh}: «{c['denominazione']}» già corretta in precedenza" if gia
                  else f"  !! {sh}: «{c['denominazione']}» non trovata nei JSON {pref}_*")

    # ---- 2. rigenerazione dei fogli toccati ----
    for sh in tocchi:
        if DRY: print(f"  (dry-run) rigenererei il foglio {sh}"); continue
        subprocess.run([sys.executable, os.path.join(BUILD,"add_country.py"), PREFIX[sh], sh, XLSX],
                       check=True, capture_output=True)
        print(f"  ↻ foglio {sh} rigenerato")

    # ---- 3. fogli senza JSON: correzione della cella in posto ----
    wb = load_workbook(XLSX)
    changed = False
    for c in C:
        sh = c["foglio"]
        if sh not in INPLACE: continue
        ws = wb[sh]
        hit = False
        for row in range(3, ws.max_row + 1):
            den = ws.cell(row, 1).value
            if not den or norm(den) != norm(c["denominazione"]): continue
            cell = ws.cell(row, COL[c["campo"]])
            cur = cell.value or ""
            if c.get("da") and str(cur).strip() != c["da"].strip():
                print(f"  ~ {sh}/{c['denominazione']}: valore attuale «{str(cur)[:60]}» "
                      f"≠ atteso «{c['da'][:60]}» — SALTATA"); hit = True; break
            if not DRY: cell.value = c["a"]
            print(f"  ✔ {sh}/{str(den)[:32]} · {c['campo']}: «{str(cur)[:45]}» → «{str(c['a'])[:45]}»")
            changed = True; hit = True; break
        if not hit:
            gia = c["campo"] == "denominazione" and any(
                norm(ws.cell(rr, 1).value) == norm(c["a"]) for rr in range(3, ws.max_row + 1))
            print(f"  = {sh}: «{c['denominazione']}» già corretta in precedenza" if gia
                  else f"  !! {sh}: «{c['denominazione']}» non trovata nel foglio")

    # ---- 4. ripristino dell'ordine dei fogli (add_country.py riaccoda in fondo) ----
    if not DRY and (changed or tocchi):
        wb._sheets = [wb[n] for n in ORDINE]
        wb.save(XLSX)
        print(f"\n↻ ordine fogli ripristinato: {load_workbook(XLSX).sheetnames}")
    elif DRY:
        print("\n(dry-run: nessuna scrittura)")

if __name__ == "__main__":
    main()

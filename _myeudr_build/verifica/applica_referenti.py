#!/usr/bin/env python3
"""
Applica al workbook i referenti e i ruoli verificati in Fase B.

Regole di sicurezza:
  · NON svuota mai un campo già valorizzato: se la proposta riguarda solo il ruolo,
    il nome resta quello che c'era.
  · Aggiorna Referente e/o Ruolo solo dove la proposta porta un valore nuovo.
  · Per i fogli con JSON di build (DK/SE/NL/BE/AT) corregge il JSON e rigenera;
    per IT/DE/FI corregge la cella in posto (non hanno JSON di origine).
  · Ripristina l'ordine dei fogli alla fine.

Uso:  python applica_referenti.py [--dry-run]
"""
import glob, json, os, re, subprocess, sys, unicodedata

HERE = os.path.dirname(os.path.abspath(__file__))
BUILD = os.path.dirname(HERE); REPO = os.path.dirname(BUILD)
XLSX = os.path.join(REPO, "MyEUDR_Lead_Mapping.xlsx")
ORDINE = ["Italia","Germania","Finlandia","Danimarca","Svezia","Olanda","Belgio","Austria"]
PREFIX = {"Danimarca":"dk","Svezia":"se","Olanda":"nl","Belgio":"be","Austria":"at"}
INPLACE = {"Italia","Germania","Finlandia"}
DRY = "--dry-run" in sys.argv

def fold(x):
    x = (x or "").lower().replace("ø","o").replace("æ","ae").replace("å","a").replace("ö","o").replace("ä","a")
    x = unicodedata.normalize("NFKD", x)
    x = "".join(c for c in x if not unicodedata.combining(c))
    x = re.sub(r"[^a-z0-9]", "", re.sub(r"\(.*?\)", "", x))
    # traslitterazioni tedesche/nordiche equivalenti: oe/ae/ue == o/a/u
    return x.replace("oe", "o").replace("ae", "a").replace("ue", "u")

def match(a, b):
    fa, fb = fold(a), fold(b)
    if fa == fb: return True
    if len(fa) >= 12 and len(fb) >= 12 and (fa.startswith(fb[:12]) or fb.startswith(fa[:12])): return True
    return sorted(fa) == sorted(fb)          # nomi con termini invertiti

def main():
    from openpyxl import load_workbook
    C = json.load(open(os.path.join(HERE, "correzioni_referenti.json"), encoding="utf-8"))
    app = skip = notfound = 0
    tocchi = set()

    # --- fogli con JSON di build ---
    for c in C:
        sh = c["foglio"]
        if sh in INPLACE: continue
        done = False
        for fp in sorted(glob.glob(os.path.join(BUILD, f"{PREFIX[sh]}_*.json"))):
            data = json.load(open(fp, encoding="utf-8")); ch = False
            for r in data:
                if not match(r.get("denominazione"), c["denominazione"]): continue
                done = True
                for k, v in (("referente", c["referente"]), ("ruolo", c["ruolo"])):
                    if not v: continue                       # niente da proporre: non tocco
                    cur = (r.get(k) or "").strip()
                    if fold(cur) == fold(v): skip += 1; continue
                    print(f"  ✔ {sh}/{r['denominazione'][:30]:32s} {k:9s} «{cur[:28]}» → «{v[:34]}»")
                    r[k] = v; ch = True; app += 1
                break
            if ch and not DRY:
                json.dump(data, open(fp, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
                tocchi.add(sh)
            if done: break
        if not done:
            print(f"  !! {sh}: «{c['denominazione']}» non trovata"); notfound += 1

    for sh in sorted(tocchi):
        if DRY: continue
        subprocess.run([sys.executable, os.path.join(BUILD, "add_country.py"), PREFIX[sh], sh, XLSX],
                       check=True, capture_output=True)
        print(f"  ↻ foglio {sh} rigenerato")

    # --- fogli senza JSON: cella in posto ---
    wb = load_workbook(XLSX); changed = False
    COL = {"referente": 4, "ruolo": 5}
    for c in C:
        sh = c["foglio"]
        if sh not in INPLACE: continue
        ws = wb[sh]; hit = False
        for row in range(3, ws.max_row + 1):
            den = ws.cell(row, 1).value
            if not den or not match(den, c["denominazione"]): continue
            hit = True
            for k, v in (("referente", c["referente"]), ("ruolo", c["ruolo"])):
                if not v: continue
                cell = ws.cell(row, COL[k]); cur = (cell.value or "").strip()
                if fold(cur) == fold(v): skip += 1; continue
                print(f"  ✔ {sh}/{str(den)[:30]:32s} {k:9s} «{cur[:28]}» → «{v[:34]}»")
                if not DRY: cell.value = v
                changed = True; app += 1
            break
        if not hit:
            print(f"  !! {sh}: «{c['denominazione']}» non trovata"); notfound += 1

    if not DRY and (changed or tocchi):
        wb._sheets = [wb[n] for n in ORDINE]; wb.save(XLSX)
        print(f"\n↻ ordine fogli ripristinato")
    print(f"\napplicate {app} · già uguali {skip} · non trovate {notfound}" + ("  (DRY-RUN)" if DRY else ""))

if __name__ == "__main__":
    main()
